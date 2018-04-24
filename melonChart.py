import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime

def get_html(url):
    _html = ""
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36',
                'Referer': 'http://www.melon.com/chart/index.htm'}
    resp = requests.get(url, headers=headers)

    if resp.status_code == 200:
        _html = resp.text

    return _html

def get_info_list(html):
    soup = BeautifulSoup(html, 'html.parser')

    tag_list_title = []
    str__list_singer = []
    tag_list_album = []
    str__list_like = []

    tr_tags = soup.tbody.find_all('tr')
    del tr_tags[0]                       # 맞춤 추천 tr태그 삭제

    for tr_tag in tr_tags:

        tags = tr_tag.find_all('div', {'class' : 'wrap_song_info'})     # title, singer, album info
        tags.extend(tr_tag.find_all('span', {'class' : 'cnt'}))         # like info

        if tags:
            ## title scrap
            tag_sub_list_title = tags[0].find(class_='ellipsis rank01').find_all('a')
            tag_list_title.extend(tag_sub_list_title)

            ## singer scrap
            tag_sub_list_singer = tags[0].find(class_='checkEllipsis').find_all('a')
            if len(tag_sub_list_singer) >= 2:

                singer = ""
                for i in range(len(tag_sub_list_singer)):
                    singer += str(tag_sub_list_singer[i].text)

                    if(i+1 == len(tag_sub_list_singer)):
                        break

                    singer += ', '
                str__list_singer.append(singer)

            else:
                str__list_singer.append(tag_sub_list_singer[0].text)

            ## album scrap
            tag_sub_list_album = tags[1].find(class_='ellipsis rank03').find_all('a')
            tag_list_album.extend(tag_sub_list_album)

            # like num
            str__sub_list_like = tags[2].text.split('\n')
            del str__sub_list_like[:2]
            str__list_like.append(str__sub_list_like[0])

    return tag_list_title, str__list_singer, tag_list_album, str__list_like


####################

# driver = webdriver.PhantomJS(executable_path='/Project/03_Src/python/melonChart/drivers/phantomjs')
driver = webdriver.Firefox(executable_path='/Project/03_Src/python/melonChart/drivers/geckodriver')

driver.get('http://www.melon.com/chart/day/index.htm')

html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

driver.close()

###################

# 차트 날짜
date = soup.find(class_='calendar_prid').find(class_='year').text
date = date.split('.')
date = "".join(date)

# 현재 날짜
n = datetime.now()
now = n.strftime('%Y%m%d')

my_file_path = "./excel/melonChart_" + now + ".xlsx"

# 파일 존재 확인
if Path(my_file_path).is_file():
    wb = load_workbook(my_file_path)    # 엑셀파일 로드
    ws = wb.active                      # 첫번째 시트
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "melonChart_day"

info_list = get_info_list(html)     # tuple형태로 정보를 가져옴

# column name
ws.cell(row=1, column=1).value = 'ranking'
ws.cell(row=1, column=2).value = 'title'
ws.cell(row=1, column=3).value = 'singer'
ws.cell(row=1, column=4).value = 'album'
ws.cell(row=1, column=5).value = 'like'
ws.cell(row=1, column=6).value = 'date'

for idx, tag in enumerate(info_list[0], 2):
    ws.cell(row=idx, column=1).value = idx - 1
    ws.cell(row=idx, column=2).value = tag.text
print('idx, title ok')

for idx, str_ in enumerate(info_list[1], 2):
    ws.cell(row=idx, column=3).value = str_
print('singer ok')

for idx, tag in enumerate(info_list[2], 2):
    ws.cell(row=idx, column=4).value = tag.text
print('album ok')

for idx, str_ in enumerate(info_list[3], 2):
    ws.cell(row=idx, column=5).value = str_
print('list number ok')

for rownum in range(2, 102):
    ws.cell(row=rownum, column=6).value = date
print('date ok')


wb.save("./excel/melonChart_" + now + ".xlsx")
wb.close()

