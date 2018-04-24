from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from pathlib import Path
from dateutil.parser import parse
from datetime import datetime, timedelta
import time


def get_info_list(html):
    soup = BeautifulSoup(html, 'html.parser')

    tag_list_title = []
    str__list_singer = []
    tag_list_album = []
    str__list_like = []

    tr_tags = soup.tbody.find_all('tr')
    del tr_tags[0]                       # 맞춤 추천 tr태그 삭제

    for tr_tag in tr_tags:

        tags = tr_tag.find_all('div', {'class' : 'wrap_song_info'})     # title, singer, album info
        tags.extend(tr_tag.find_all('span', {'class' : 'cnt'}))         # like info

        if tags:
            # title scrap
            tag_sub_list_title = tags[0].find(class_='ellipsis rank01').find_all('a')
            tag_list_title.extend(tag_sub_list_title)

            # singer scrap
            tag_sub_list_singer = tags[0].find(class_='checkEllipsis').find_all('a')
            if len(tag_sub_list_singer) >= 2:

                singer = ""
                length = len(tag_sub_list_singer)
                for i in range(length):
                    singer += str(tag_sub_list_singer[i].text)

                    if(i+1 == length):
                        break

                    singer += ', '
                str__list_singer.append(singer)

            else:
                str__list_singer.append(tag_sub_list_singer[0].text)

            # album scrap
            tag_sub_list_album = tags[1].find(class_='ellipsis rank03').find_all('a')
            tag_list_album.extend(tag_sub_list_album)

            # like num
            str__sub_list_like = tags[2].text.split('\n')
            del str__sub_list_like[:2]
            str__list_like.append(str__sub_list_like[0])

    return tag_list_title, str__list_singer, tag_list_album, str__list_like

###################################
# driver = webdriver.PhantomJS(executable_path='/Project/03_Src/python/melonChart/drivers/phantomjs')
driver = webdriver.Firefox(executable_path='/Project/03_Src/python/melonChart/drivers/geckodriver')
###################################

n = datetime.now()
now = n.strftime('%Y%m%d')
my_file_path = "./excel/melonChartWeek_" + now + ".xlsx"

# 파일 존재 확인
if Path(my_file_path).is_file():
    wb = load_workbook(my_file_path)  # 엑셀파일 로드
    ws = wb.active  # 첫번째 시트

else:
    wb = Workbook()
    ws = wb.active
    ws.title = "melonChart_week"    # sheet name
###################################


url = 'http://www.melon.com/chart/week/index.htm'
cnt = 0

for i in range(12):

    driver.get(url)

    time.sleep(1)

    html = driver.page_source

    # 현재 url 에서 startDay, endDay 가져오기
    startDay = driver.current_url.split('&')[1].split('=')[1]
    endDay = driver.current_url.split('&')[2].split('=')[1]

    ###################

    info_list = get_info_list(html)     # tuple형태로 정보를 가져옴

    if i == 0:
        # column name
        ws.cell(row=1, column=1).value = 'ranking'
        ws.cell(row=1, column=2).value = 'title'
        ws.cell(row=1, column=3).value = 'singer'
        ws.cell(row=1, column=4).value = 'album'
        ws.cell(row=1, column=5).value = 'like'
        ws.cell(row=1, column=6).value = 'startDay'
        ws.cell(row=1, column=7).value = 'endDay'

    for idx, tag in enumerate(info_list[0], 2):
        ws.cell(row=idx + cnt, column=1).value = idx-1
        ws.cell(row=idx + cnt, column=2).value = tag.text
    print('idx, title ok')

    for idx, str_ in enumerate(info_list[1], 2):
        ws.cell(row=idx + cnt, column=3).value = str_
    print('singer ok')

    for idx, tag in enumerate(info_list[2], 2):
        ws.cell(row=idx + cnt, column=4).value = tag.text
    print('album ok')

    for idx, str_ in enumerate(info_list[3], 2):
        ws.cell(row=idx + cnt, column=5).value = str_
    print('list number ok')

    for rownum in range(2, 102):
        ws.cell(row=rownum + cnt, column=6).value = startDay
        ws.cell(row=rownum + cnt, column=7).value = endDay
    print('startDay, endDay ok')

    cnt += 100

    #########################
    # 현재 날짜에서 7일을 빼고
    startDay = str(parse(startDay) - timedelta(days=7)).split(' ')[0].split('-')
    endDay = str(parse(endDay) - timedelta(days=7)).split(' ')[0].split('-')

    # 다시 String type으로
    startDay = "".join(startDay)
    endDay = "".join(endDay)

    url = 'http://www.melon.com/chart/week/index.htm' \
          '#params[idx]=1' + '&params[startDay]=' + startDay + '&params[endDay]=' + endDay
    #########################

wb.save("./excel/melonChartWeek_" + now + ".xlsx")

wb.close()
driver.close()










